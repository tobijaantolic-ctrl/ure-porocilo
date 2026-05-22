import streamlit as st
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re, io, math

st.set_page_config(page_title="Poročilo opravljenih ur", page_icon="📋", layout="centered")
st.title("📋 Poročilo opravljenih ur")
st.markdown("Naloži **PDF poročilo učitelja** in prejmi Excel datoteko z razčlenjenimi urami.")

# ─────────────────────────────────────────────────────────────────────────────
# PRAVILO RAZVRŠČANJA (edino merilo):
#   naziv se ZAČNE z "wau"  →  WAU  (dejanski čas, ne zaokroži)
#   naziv vsebuje "RaP"     →  RAP  (pedagoška: ceil(h/0.75))
#   vse ostalo              →  Dirka za branje (pedagoška: ceil(h/0.75))
# ─────────────────────────────────────────────────────────────────────────────

def pedagoske_ure(decimal: float) -> int:
    return max(1, math.ceil(decimal / 0.75))

def ocisti_naziv(raw: str) -> str:
    n = raw.strip()
    if n.lower().startswith("wau "):
        n = n[4:].strip()
    # Odstrani priponko razreda: " 3."  " 11."  " 0."
    n = re.sub(r'\s+\d+\.\s*$', '', n).strip()
    return n

def rap_tip(raw: str) -> str:
    m = re.search(r'(RaP\s*(?:PS|RS)|RAP)', raw, re.IGNORECASE)
    return m.group(1) if m else ocisti_naziv(raw)

# Regex — brez DOTALL da ne preskakuje vrstic
VRSTICA = re.compile(
    r'(\d{1,2}\.\s*\d{1,2}\.\s*\d{4})'   # datum
    r'\s+([^\n\r]+?)'                       # naziv + skupina (ena vrstica)
    r'\s+\d+h\s+\d+m'                       # čas (zavržemo)
    r'\s+\((\d+(?:[.,]\d+)?)\)'             # (ure decimalno)
    r'\s+(Sistemizirana|Nesistemizirana)'    # tip
)

def preberi_pdf(pdf_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    # Nekateri nazivi se v PDF-ju prelomijo v dve vrstici (npr. Gumitvist).
    # Nadaljevalna vrstica začne z razrednim znakom pred časom: "2_5 8. 0h 45m"
    # Spojimo jo s predhodno vrstico.
    import re as _re
    text = _re.sub(r'\n([\w_]+\s+\d+\.\s+\d+h\s+\d+m)', r' \1', text)
    return text

def razcleni_pdf(text: str) -> dict:
    ucitelj = "Neznano"
    m = re.search(r'Poročilo učitelja\s*\n([^\n]+)', text)
    if m:
        ucitelj = m.group(1).strip()

    obdobje = ""
    m = re.search(r'Obdobje:\s*(od\s+[\d\.\s]+do\s+[\d\.\s]+)', text)
    if m:
        obdobje = m.group(1).strip()

    rap, dirka, wau = [], [], []

    for m in VRSTICA.finditer(text):
        raw_datum, raw_naziv, ure_str, _tip = m.groups()
        datum = re.sub(r'\s+', '', raw_datum)
        ure   = float(ure_str.replace(',', '.'))
        niz   = raw_naziv.strip()

        if niz.lower().startswith("wau"):
            # ── WAU: SAMO "wau" predpona ──────────────────────────────────
            wau.append({
                "datum": datum,
                "opis":  ocisti_naziv(niz),
                "ure":   ure,
            })
        elif re.search(r'\brap\b', niz, re.IGNORECASE):
            # ── RAP: vsebuje "RaP" ────────────────────────────────────────
            rap.append({
                "datum":   datum,
                "naziv":   rap_tip(niz),
                "ure_st":  pedagoske_ure(ure),
            })
        else:
            # ── VSE OSTALO: pedagoška ura ─────────────────────────────────
            dirka.append({
                "datum":     datum,
                "naziv":     ocisti_naziv(niz),
                "dejavnost": _tip,
                "ure_st":    pedagoske_ure(ure),
            })

    return {
        "ucitelj": ucitelj,
        "obdobje": obdobje,
        "rap":   rap,
        "dirka": dirka,
        "wau":   wau,
    }

# ── Excel styling ─────────────────────────────────────────────────────────────
def _rob():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def glava(cell, text):
    cell.value = text
    cell.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill  = PatternFill("solid", start_color="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _rob()

def celica(cell, value, alt=False, fmt=None):
    cell.value = value
    cell.font  = Font(name="Arial", size=11)
    cell.fill  = PatternFill("solid", start_color="DCE6F1") if alt else PatternFill()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _rob()
    if fmt:
        cell.number_format = fmt

def skupaj(cell, value, fmt=None):
    cell.value = value
    cell.font  = Font(bold=True, name="Arial", size=11)
    cell.fill  = PatternFill("solid", start_color="FFC000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _rob()
    if fmt:
        cell.number_format = fmt

def sirine(ws, widths):
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w

def ustvari_excel(data: dict) -> bytes:
    wb    = Workbook()
    rap   = data["rap"]
    dirka = data["dirka"]
    wau   = data["wau"]

    # Sheet 1 — RAP
    ws1 = wb.active; ws1.title = "RAP_RaP_PS"
    for c, h in enumerate(["Datum", "Naziv", "Ure (upoštevano)"], 1):
        glava(ws1.cell(1, c), h)
    for i, r in enumerate(rap):
        celica(ws1.cell(i+2, 1), r["datum"],  i%2)
        celica(ws1.cell(i+2, 2), r["naziv"],  i%2)
        celica(ws1.cell(i+2, 3), r["ure_st"], i%2, "0")
    tr1 = len(rap) + 2
    ws1.merge_cells(f"A{tr1}:B{tr1}")
    skupaj(ws1.cell(tr1, 1), "SKUPAJ")
    skupaj(ws1.cell(tr1, 3), f"=SUM(C2:C{tr1-1})", "0")
    sirine(ws1, [14, 22, 20])

    # Sheet 2 — Dirka za branje (vse ostale pedagoške)
    ws2 = wb.create_sheet("Dirka_za_branje")
    for c, h in enumerate(["Datum", "Naziv", "Dejavnost", "Ure (upoštevano)"], 1):
        glava(ws2.cell(1, c), h)
    for i, r in enumerate(dirka):
        celica(ws2.cell(i+2, 1), r["datum"],     i%2)
        celica(ws2.cell(i+2, 2), r["naziv"],     i%2)
        celica(ws2.cell(i+2, 3), r["dejavnost"], i%2)
        celica(ws2.cell(i+2, 4), r["ure_st"],    i%2, "0")
    tr2 = len(dirka) + 2
    ws2.merge_cells(f"A{tr2}:C{tr2}")
    skupaj(ws2.cell(tr2, 1), "SKUPAJ")
    skupaj(ws2.cell(tr2, 4), f"=SUM(D2:D{tr2-1})", "0")
    sirine(ws2, [14, 36, 18, 20])

    # Sheet 3 — WAU (samo "wau" predpona, dejanski čas)
    ws3 = wb.create_sheet("WAU")
    for c, h in enumerate(["Datum", "Opis", "Ure (dejanski čas)"], 1):
        glava(ws3.cell(1, c), h)
    for i, r in enumerate(wau):
        celica(ws3.cell(i+2, 1), r["datum"], i%2)
        celica(ws3.cell(i+2, 2), r["opis"],  i%2)
        celica(ws3.cell(i+2, 3), r["ure"],   i%2, "0.00")
    tr3 = len(wau) + 2
    ws3.merge_cells(f"A{tr3}:B{tr3}")
    skupaj(ws3.cell(tr3, 1), "SKUPAJ")
    skupaj(ws3.cell(tr3, 3), f"=SUM(C2:C{tr3-1})", "0.00")
    sirine(ws3, [14, 32, 22])

    # Sheet 4 — Poročilo o urah
    ws4 = wb.create_sheet("Poročilo o urah")
    for c, h in enumerate(["Tabela", "Ure"], 1):
        glava(ws4.cell(1, c), h)
    for i, (lbl, fml) in enumerate([
        ("RAP / RaP PS",    f"=RAP_RaP_PS!C{tr1}"),
        ("Dirka za branje", f"=Dirka_za_branje!D{tr2}"),
        ("WAU",             f"=WAU!C{tr3}"),
    ]):
        celica(ws4.cell(i+2, 1), lbl, i%2)
        celica(ws4.cell(i+2, 2), fml, i%2, "0.00")
    skupaj(ws4.cell(5, 1), "SKUPAJ")
    skupaj(ws4.cell(5, 2), "=SUM(B2:B4)", "0.00")
    sirine(ws4, [24, 14])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── UI ────────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("📄 Naloži PDF poročilo", type=["pdf"])

if uploaded:
    st.info(f"Datoteka: **{uploaded.name}**")
    if st.button("⚙️ Obdelaj in ustvari Excel", type="primary"):
        with st.spinner("Berem in analiziram PDF..."):
            text = preberi_pdf(uploaded.read())
            data = razcleni_pdf(text)

        rap_h   = sum(r["ure_st"] for r in data["rap"])
        dirka_h = sum(r["ure_st"] for r in data["dirka"])
        wau_h   = sum(r["ure"]   for r in data["wau"])
        total   = len(data["rap"]) + len(data["dirka"]) + len(data["wau"])

        if total == 0:
            st.error("Ni aktivnosti. Surovo besedilo za debug:")
            st.text_area("", text[:3000], height=200)
        else:
            excel_bytes = ustvari_excel(data)
            st.success("✅ Excel je pripravljen!")

            c1, c2 = st.columns(2)
            c1.markdown(f"**Učitelj:** {data['ucitelj']}")
            c2.markdown(f"**Obdobje:** {data['obdobje']}")

            c3, c4, c5 = st.columns(3)
            c3.metric("RAP / RaP PS",    f"{rap_h} ur",    f"{len(data['rap'])} vnosov")
            c4.metric("Dirka za branje", f"{dirka_h} ur",  f"{len(data['dirka'])} vnosov")
            c5.metric("WAU",             f"{wau_h:.2f} ur", f"{len(data['wau'])} vnosov")

            safe = re.sub(r'[^\w]', '_', data["ucitelj"])
            st.download_button(
                "⬇️ Prenesi Excel", excel_bytes,
                file_name=f"porocilo_{safe}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

st.markdown("---")
st.caption("Šmarje-Sap · Poročilo opravljenih ur · v3.0")
